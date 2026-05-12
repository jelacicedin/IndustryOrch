# Copyright (C) 2026  Edin Jelacic — AGPL-3.0-or-later
"""File converter for industrial documents into structured markdown.

Supported formats: PDF, XLSX, CSV, MD, TXT.

PDFs use PyMuPDF for fast text + image extraction and pdfplumber for accurate
table extraction. Images are described via an Ollama vision model when
VISION_MODEL is configured; otherwise their presence is noted as a placeholder.
"""

from __future__ import annotations

import argparse
import base64
import datetime
import logging
import os
import re
from pathlib import Path
from typing import Any, Callable

logger = logging.getLogger("industryorch-ingest")

import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

VISION_MODEL = os.environ.get("VISION_MODEL", "").strip()
OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://localhost:11434")

# PDF extraction mode:
#   "hybrid" (default) — PyMuPDF text + pdfplumber tables + optional image description
#   "vision"           — render each page to PNG, send to vision LLM for full extraction
PDF_EXTRACT_MODE = os.environ.get("PDF_EXTRACT_MODE", "hybrid").lower()

# Resolution for vision-mode page rendering (144 DPI = 2× PyMuPDF base of 72 DPI)
_VISION_DPI_SCALE = float(os.environ.get("PDF_VISION_DPI_SCALE", "2.0"))

# Skip images smaller than this (filters out icons / decorative elements)
_MIN_IMG_WIDTH = 100
_MIN_IMG_HEIGHT = 100

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(r"\b(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\b")
_REVISION_RE = re.compile(r"\b(v?\d+\.\d+(?:\.\d+)?)\b")


def _find_first_dates(text: str) -> list[datetime.date]:
    dates: list[datetime.date] = []
    for m in _DATE_RE.finditer(text):
        raw = m.group(1)
        for fmt in (
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%Y-%m-%d",
            "%m/%d/%y",
            "%d-%m-%y",
        ):
            try:
                dates.append(datetime.datetime.strptime(raw, fmt).date())
                break
            except ValueError:
                continue
    return sorted(dates)


def _find_first_revision(text: str) -> str | None:
    m = _REVISION_RE.search(text)
    return m.group(0) if m else None


def _escape_md_cell(val: Any) -> str:
    return str(val).replace("|", "\\|").replace("\n", " ").strip()


def _dataframe_to_md(df: pd.DataFrame) -> str:
    headers = df.columns.astype(str).tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for _, row in df.iterrows():
        cells = " | ".join(_escape_md_cell(v) for v in row.values)
        lines.append("| " + cells + " |")
    return "\n".join(lines)


def _plumber_table_to_md(table: list[list]) -> str:
    """Convert a pdfplumber table to markdown."""
    if not table or not table[0]:
        return ""
    header = [_escape_md_cell(c) for c in table[0]]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in table[1:]:
        cells = [_escape_md_cell(c) for c in row]
        cells += [""] * max(0, len(header) - len(cells))
        lines.append("| " + " | ".join(cells[: len(header)]) + " |")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Vision model image description
# ---------------------------------------------------------------------------


def _describe_image(
    img_bytes: bytes, keep_alive: str | int = "5m"
) -> str | None:
    """Describe an image using the configured Ollama vision model."""
    if not VISION_MODEL:
        return None
    try:
        import ollama

        resp = ollama.chat(
            model=VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": (
                        "Describe this technical diagram, chart, or image "
                        "concisely. Focus on labels, measurements, data "
                        "values, and what it depicts."
                    ),
                    "images": [base64.b64encode(img_bytes).decode()],
                }
            ],
            keep_alive=keep_alive,
        )
        return resp.message.content.strip()
    except Exception:
        return None


def _evict_vision_model() -> None:
    """Evict the vision model from VRAM after all images are processed.

    Prevents the vision model and the graph extraction / generation model
    from occupying VRAM simultaneously.
    """
    if not VISION_MODEL:
        return
    try:
        import ollama
        ollama.generate(model=VISION_MODEL, prompt="", keep_alive=0)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def _convert_pdf_hybrid(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Extract structured markdown from a PDF.

    Uses PyMuPDF for fast text + image extraction, pdfplumber for tables.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError("Install PyMuPDF: pip install pymupdf")

    import pdfplumber

    path = Path(file_path)
    title = path.stem

    doc = fitz.open(file_path)
    page_count = len(doc)
    md_parts: list[str] = []
    images_described = False

    with pdfplumber.open(file_path) as pdf:
        for page_idx in range(page_count):
            if progress_callback:
                progress_callback(
                    f"Converting page {page_idx + 1}/{page_count} …"
                )

            fitz_page = doc[page_idx]
            plumb_page = pdf.pages[page_idx]

            if page_count > 1:
                heading = f"## Page {page_idx + 1}"
            else:
                heading = f"# {title}"
            section: list[str] = [heading]

            # Text via PyMuPDF — faster than pdfplumber for plain text
            text = fitz_page.get_text("text") or ""
            if text.strip():
                section.append(f"### Text\n\n{text.strip()}")

            # Tables via pdfplumber — more accurate for complex layouts
            tables = plumb_page.extract_tables()
            if tables:
                valid = [(i, t) for i, t in enumerate(tables, 1) if t]
                for t_idx, table in valid:
                    md = _plumber_table_to_md(table)
                    if md:
                        label = (
                            f"### Table {t_idx}"
                            if len(valid) > 1
                            else "### Table"
                        )
                        section.append(f"{label}\n\n{md}")

            # Images via PyMuPDF
            img_refs = fitz_page.get_images(full=True)
            img_parts: list[str] = []
            for img_num, img_ref in enumerate(img_refs, start=1):
                xref = img_ref[0]
                try:
                    info = doc.extract_image(xref)
                    w, h = info.get("width", 0), info.get("height", 0)
                    if w < _MIN_IMG_WIDTH or h < _MIN_IMG_HEIGHT:
                        continue
                    desc = _describe_image(info["image"])
                    if desc:
                        images_described = True
                        img_parts.append(
                            f"**[Figure {img_num}]** ({w}×{h}px): {desc}"
                        )
                    else:
                        hint = (
                            " Set VISION_MODEL to enable descriptions."
                            if not VISION_MODEL
                            else ""
                        )
                        img_parts.append(
                            f"**[Figure {img_num}]** ({w}×{h}px): "
                            f"*Technical diagram/image present.*{hint}"
                        )
                except Exception:
                    continue

            if img_parts:
                section.append("### Figures\n\n" + "\n\n".join(img_parts))

            md_parts.append("\n\n".join(section))
            md_parts.append("---")

    doc.close()

    # Free vision model VRAM before graph extraction / generation loads
    if images_described:
        _evict_vision_model()

    full_md = "\n\n".join(md_parts)
    first_dates = _find_first_dates(full_md)

    return full_md, {
        "title": title,
        "detected_date": first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(full_md),
        "page_count": page_count,
    }


# ---------------------------------------------------------------------------
# PDF — vision extraction mode
# ---------------------------------------------------------------------------

_VISION_EXTRACTION_PROMPT = """\
This is page {page_num} of {page_count} from an industrial technical document.
Extract ALL content as structured markdown. Rules:
- Use ##, ### for section headings found on the page
- Convert EVERY table to a markdown table (| col | col |\\n|---|---|\\n| val | val |)
- For figures, charts, or diagrams write one line: **[Figure]** brief description
- Copy all numbers, units, part numbers, codes, and identifiers exactly as shown
- Do not add commentary — output only the extracted content\
"""


def _extract_page_vision(
    img_bytes: bytes,
    page_num: int,
    page_count: int,
    keep_alive: str | int = "5m",
) -> str | None:
    """Send a rendered page image to the vision LLM and return markdown."""
    if not VISION_MODEL:
        return None
    try:
        import ollama

        prompt = _VISION_EXTRACTION_PROMPT.format(
            page_num=page_num, page_count=page_count
        )
        resp = ollama.chat(
            model=VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": prompt,
                    "images": [base64.b64encode(img_bytes).decode()],
                }
            ],
            keep_alive=keep_alive,
        )
        text = (resp.message.content or "").strip()
        return text if text else None
    except Exception as exc:
        logger.warning("Vision extraction failed for page %d: %s", page_num, exc)
        return None


def _convert_pdf_vision(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Extract markdown from a PDF by rendering pages as images and using a vision LLM.

    Each page is rendered to PNG at _VISION_DPI_SCALE × 72 DPI (default 144 DPI)
    and sent to VISION_MODEL with a structured extraction prompt. Falls back to
    the hybrid extractor for any page where the vision model returns nothing.
    """
    try:
        import fitz
    except ImportError:
        raise ImportError("Install PyMuPDF: pip install pymupdf")

    path = Path(file_path)
    title = path.stem
    doc = fitz.open(file_path)
    page_count = len(doc)
    md_parts: list[str] = []
    mat = fitz.Matrix(_VISION_DPI_SCALE, _VISION_DPI_SCALE)

    logger.info(
        "PDF vision extraction: %d page(s), model=%s, scale=%.1fx",
        page_count, VISION_MODEL, _VISION_DPI_SCALE,
    )

    for page_idx in range(page_count):
        if progress_callback:
            progress_callback(
                f"Extracting page {page_idx + 1}/{page_count} (vision) …"
            )

        fitz_page = doc[page_idx]
        heading = f"## Page {page_idx + 1}" if page_count > 1 else f"# {title}"

        is_last = page_idx == page_count - 1
        keep_alive: str | int = 0 if is_last else "5m"

        # Render page to PNG bytes
        pix = fitz_page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")

        md = _extract_page_vision(
            img_bytes, page_idx + 1, page_count, keep_alive=keep_alive
        )

        if md:
            md_parts.append(f"{heading}\n\n{md}")
        else:
            # Vision returned nothing — fall back to simple text extraction
            logger.warning(
                "Vision returned nothing for page %d — using text fallback",
                page_idx + 1,
            )
            text = fitz_page.get_text("text") or ""
            if text.strip():
                md_parts.append(f"{heading}\n\n{text.strip()}")

        md_parts.append("---")

    doc.close()

    full_md = "\n\n".join(md_parts)
    first_dates = _find_first_dates(full_md)

    return full_md, {
        "title": title,
        "detected_date": first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(full_md),
        "page_count": page_count,
        "extract_mode": "vision",
    }


def _convert_pdf(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Dispatch to vision or hybrid PDF extraction based on PDF_EXTRACT_MODE."""
    if PDF_EXTRACT_MODE == "vision":
        if not VISION_MODEL:
            logger.warning(
                "PDF_EXTRACT_MODE=vision but VISION_MODEL is not set "
                "— falling back to hybrid extraction"
            )
        else:
            return _convert_pdf_vision(file_path, progress_callback)
    return _convert_pdf_hybrid(file_path, progress_callback)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _convert_xlsx(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Convert XLSX workbook to structured markdown.

    Each sheet becomes a ## Sheet: <name> section.
    Merged cells are handled by column-wise fill-down.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    md_parts: list[str] = []
    sheet_names: list[str] = []
    total = len(wb.sheetnames)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        if progress_callback:
            progress_callback(
                f"Converting sheet {sheet_idx}/{total}: {sheet_name} …"
            )

        ws = wb[sheet_name]
        sheet_names.append(sheet_name)

        num_rows = ws.max_row or 0
        num_cols = ws.max_column or 0
        if num_rows == 0 or num_cols == 0:
            continue

        raw: dict[tuple[int, int], Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                raw[(cell.row - 1, cell.column - 1)] = cell.value

        fill_down: dict[int, Any] = {}
        for col in range(num_cols):
            for row in range(num_rows):
                val = raw.get((row, col))
                if val is not None:
                    fill_down[col] = val
                elif col in fill_down:
                    raw[(row, col)] = fill_down[col]

        rows_list = [
            tuple(raw.get((row, col)) for col in range(num_cols))
            for row in range(num_rows)
            if any(raw.get((row, col)) is not None for col in range(num_cols))
        ]
        if not rows_list:
            continue

        md_parts.append(f"## Sheet: {sheet_name}")
        header = [_escape_md_cell(c) for c in rows_list[0]]
        md_parts.append("| " + " | ".join(header) + " |")
        md_parts.append("| " + " | ".join("---" for _ in header) + " |")
        for data_row in rows_list[1:]:
            cells = [_escape_md_cell(c) for c in list(data_row)[: len(header)]]
            cells += [""] * (len(header) - len(cells))
            md_parts.append("| " + " | ".join(cells) + " |")
        md_parts.append("")

    wb.close()
    text = "\n".join(md_parts)
    first_dates = _find_first_dates(text)

    return text, {
        "title": Path(file_path).stem,
        "detected_date": first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(text),
        "sheet_names": sheet_names,
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _convert_csv(file_path: str) -> tuple[str, dict[str, Any]]:
    """Convert CSV to markdown table, splitting large files into chunks."""
    df = pd.read_csv(file_path)
    chunk_size = 100
    md_parts: list[str] = []

    if len(df) <= chunk_size:
        md_parts.append(f"# {Path(file_path).stem}")
        md_parts.append(_dataframe_to_md(df))
    else:
        n = (len(df) + chunk_size - 1) // chunk_size
        for i in range(n):
            start, end = i * chunk_size, min((i + 1) * chunk_size, len(df))
            md_parts.append(f"## Chunk {i + 1}/{n} (rows {start + 1}–{end})")
            md_parts.append(_dataframe_to_md(df.iloc[start:end]))
            md_parts.append("")

    text = "\n\n".join(md_parts)
    first_dates = _find_first_dates(text)

    return text, {
        "title": Path(file_path).stem,
        "detected_date": first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(text),
        "row_count": len(df),
        "column_count": len(df.columns),
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def convert_file(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Convert a file to structured markdown suitable for chunking.

    Parameters
    ----------
    file_path         : Path to the file (.pdf, .xlsx, .csv, .md, .txt).
    progress_callback : Optional callable(str) for per-page/sheet updates.

    Returns
    -------
    tuple[str, dict]
        (markdown_string, metadata_dict). Metadata always includes
        *title*, *detected_date*, *detected_revision*.
    """
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".pdf":
        return _convert_pdf(file_path, progress_callback=progress_callback)
    elif ext in (".xlsx", ".xls"):
        return _convert_xlsx(file_path, progress_callback=progress_callback)
    elif ext == ".csv":
        return _convert_csv(file_path)
    elif ext in (".md", ".txt"):
        content = path.read_text(encoding="utf-8", errors="replace")
        return content, {
            "title": path.stem,
            "detected_date": None,
            "detected_revision": None,
            "file_type": ext.lstrip("."),
        }
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert industrial documents to structured markdown"
    )
    parser.add_argument("--file", required=True, help="Path to the input file")
    args = parser.parse_args()

    import json

    markdown, metadata = convert_file(
        args.file, progress_callback=lambda m: print(f"  {m}")
    )
    print("---METADATA---")
    print(json.dumps(metadata, indent=2, default=str))
    print("---MARKDOWN---")
    print(markdown)


if __name__ == "__main__":
    main()
